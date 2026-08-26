#!/usr/bin/env python3
"""Export LOTO placards to an Excel workbook that mirrors the printed PDF.

The workbook renders one page per placard per language, laid out band for
band against `apps/web/lib/pdfPlacard.ts` — yellow title band, blue
equipment bar, red warning block, purpose + application columns, colour
codes grid, photo slots, energy-steps table, removal process, signature
bar. Alongside those pages it carries flat Equipment and Energy Steps
sheets for filtering, and a Summary sheet of per-department counts.

Two subcommands, because fetching and rendering fail for different
reasons and are usefully retried apart:

    fetch  --out data/            reads Supabase, caches JSON + photos
    build  --data data/ --out placards.xlsx    renders the workbook

`build` never touches the network: it embeds whatever photos `fetch`
cached and falls back to the PDF's own "-- No photo --" slot for the
rest, so an export still succeeds where egress is closed.

Environment (fetch only):
    SUPABASE_URL                 https://<ref>.supabase.co
    SUPABASE_SERVICE_ROLE_KEY    service-role JWT (RLS-exempt; keep secret)
    SUPABASE_TENANT_NUMBER       4-digit tenant, default 0001
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
import urllib.parse
import urllib.request
from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break
from openpyxl.worksheet.page import PageMargins

# ── Placard constants, mirrored from the TypeScript sources ─────────────────
# energyCodes.ts — code registry. 'N' is an app-side sentinel the placard
# legend filters out, so it is last and excluded from the grid.
ENERGY_CODES: list[tuple[str, str, str, str, str]] = [
    # code, label_en,          label_es,          fill,      text
    ("E",  "Electrical",       "Eléctrico",       "BF1414", "FFFFFF"),
    ("G",  "Gas",              "Gas",             "FFD900", "1A1A1A"),
    ("H",  "Hydraulic",        "Hidráulico",      "A67B5B", "FFFFFF"),
    ("P",  "Pneumatic",        "Neumático",       "1478C7", "FFFFFF"),
    ("M",  "Mechanical",       "Mecánico",        "7F4DB3", "FFFFFF"),
    ("T",  "Thermal",          "Térmico",         "000000", "FFFFFF"),
    ("W",  "Water",            "Agua",            "33993A", "FFFFFF"),
    ("S",  "Steam",            "Vapor",           "E07B00", "FFFFFF"),
    ("V",  "Valve",            "Válvula",         "888888", "FFFFFF"),
    ("CG", "Compressed Gas",   "Gas Comprimido",  "0E8A8A", "FFFFFF"),
    ("CP", "Control Panel",    "Panel Control",   "FFFFFF", "1A1A1A"),
    ("GR", "Gravity",          "Gravedad",        "8B0A1A", "FFFFFF"),
    ("N",  "None",             "Ninguno",         "555555", "FFFFFF"),
]
ENERGY_ALIASES = {"O": "M", "OG": "CG"}
_BY_CODE = {c[0]: c for c in ENERGY_CODES}

# lotoProcedureValidation.ts — OSHA documentation order. The placard groups
# steps by phase in this order, then by sequence_order within each phase.
STEP_PHASE_ORDER = [
    "shutdown",
    "isolate",
    "release_stored_energy",
    "lockout",
    "verify_zero_energy",
]
PHASE_LABELS = {
    "en": {
        "shutdown": "SHUTDOWN",
        "isolate": "ISOLATE",
        "release_stored_energy": "RELEASE STORED ENERGY",
        "lockout": "LOCKOUT",
        "verify_zero_energy": "VERIFY ZERO ENERGY",
    },
    "es": {
        "shutdown": "APAGADO",
        "isolate": "AISLAR",
        "release_stored_energy": "LIBERAR ENERGIA ALMACENADA",
        "lockout": "BLOQUEO",
        "verify_zero_energy": "VERIFICAR ENERGIA CERO",
    },
}

# placardText.ts — every fixed string the placard prints.
TEXT = {
    "title": {
        "en": "LOCKOUT/TAGOUT PROCEDURE",
        "es": "PROCEDIMIENTO DE BLOQUEO/ETIQUETADO",
    },
    "equipment_label": {"en": "EQUIPMENT:", "es": "EQUIPO:"},
    "warning_header": {
        "en": "KEEP OUT! HAZARDOUS VOLTAGE AND MOVING PARTS",
        "es": "¡MANTÉNGASE ALEJADO! VOLTAJE PELIGROSO Y PIEZAS EN MOVIMIENTO",
    },
    "warning_fallback": {
        "en": "This equipment must be locked out and tagged out before servicing or "
              "maintenance. Follow the procedure below to isolate all energy sources.",
        "es": "Este equipo debe bloquearse y etiquetarse antes de darle servicio o "
              "mantenimiento. Siga el procedimiento a continuación para aislar "
              "todas las fuentes de energía.",
    },
    "purpose_header": {"en": "PURPOSE", "es": "PROPÓSITO"},
    "purpose_body": {
        "en": "This procedure establishes the minimum requirements for lockout of "
              "energy-isolating devices. It ensures equipment is stopped, isolated "
              "from all potentially hazardous energy sources, and locked out before "
              "any servicing or maintenance activities are performed.",
        "es": "Este procedimiento establece los requisitos mínimos para el bloqueo "
              "de los dispositivos de aislamiento de energía. Garantiza que el "
              "equipo se detenga, se aísle de todas las fuentes de energía "
              "potencialmente peligrosas y se bloquee antes de realizar actividades "
              "de servicio o mantenimiento.",
    },
    "application_header": {
        "en": "LOCKOUT APPLICATION PROCESS",
        "es": "PROCESO DE APLICACIÓN DE BLOQUEO",
    },
    "application_steps": {
        "en": [
            "Communicate to all AFFECTED employees",
            "Shut down the equipment using normal stopping procedures",
            "Isolate energy sources",
            "Apply lockout devices, locks, and tags",
            "Release all stored energy",
            "Verify equipment is de-energized by attempting to start up",
            "After test, place controls in a neutral position",
        ],
        "es": [
            "Comunicar a todos los empleados AFECTADOS",
            "Apagar el equipo usando procedimientos normales de parada",
            "Aislar las fuentes de energía",
            "Aplicar dispositivos de bloqueo, candados y etiquetas",
            "Liberar toda la energía almacenada",
            "Verificar que el equipo esté desenergizado intentando arrancarlo",
            "Después de la prueba, colocar los controles en posición neutral",
        ],
    },
    "removal_header": {
        "en": "LOCKOUT REMOVAL PROCESS",
        "es": "PROCESO DE REMOCIÓN DE BLOQUEO",
    },
    "removal_steps": {
        "en": [
            "Notify all AFFECTED employees that lockout is being removed",
            "Inspect the work area to ensure tools and items have been removed",
            "Verify that all employees are clear of the equipment",
            "Verify that controls are in the neutral or off position",
            "Remove lockout devices, locks, and tags",
            "Re-energize the equipment",
            "Notify all AFFECTED employees that the equipment is back in service",
        ],
        "es": [
            "Notificar a todos los empleados AFECTADOS que se está retirando el bloqueo",
            "Inspeccionar el área de trabajo para asegurar que se hayan retirado "
            "herramientas y objetos",
            "Verificar que todos los empleados estén lejos del equipo",
            "Verificar que los controles estén en posición neutral o apagada",
            "Retirar los dispositivos de bloqueo, candados y etiquetas",
            "Volver a energizar el equipo",
            "Notificar a los empleados AFECTADOS que el equipo está de vuelta en servicio",
        ],
    },
    "color_codes_header": {"en": "COLOR CODES", "es": "CÓDIGOS DE COLOR"},
    "section_header": {
        "en": "EQUIPMENT IDENTIFICATION AND ENERGY ISOLATION PROCEDURE",
        "es": "IDENTIFICACIÓN DEL EQUIPO Y PROCEDIMIENTO DE AISLAMIENTO DE ENERGÍA",
    },
    "photo_captions": {
        "en": ("PHOTO OF EQUIPMENT", "PHOTO OF ISOLATION / DISCONNECT"),
        "es": ("FOTO DEL EQUIPO", "FOTO DE AISLAMIENTO / DESCONEXIÓN"),
    },
    "table_headers": {
        "en": ("Energy Tag & Description", "Isolation Procedure & Lockout Devices",
               "Method of Verification"),
        "es": ("Etiqueta y Descripción", "Procedimiento de Aislamiento y Dispositivos",
               "Método de Verificación"),
    },
    "signature": {
        "en": ("Signature", "Date", "Dept", "See PM Store in PT Folder"),
        "es": ("Firma", "Fecha", "Depto", "Ver PM Store en carpeta PT"),
    },
    "no_steps": {
        "en": "No energy steps defined for this equipment.",
        "es": "No hay pasos de energía definidos para este equipo.",
    },
    "no_photo": {"en": "— No photo —", "es": "— Sin foto —"},
    "print_note": {
        "en": "Spanish translation on reverse — print double-sided.",
        "es": "Traducción al inglés al reverso — imprimir a doble cara.",
    },
}

# Palette, from pdfPlacard.ts.
YELLOW_BAND = "FFD900"
BLUE_BAR = "D9E8FF"
RED_BLOCK = "BF1414"
NAVY_HEADER = "214488"
WHITE = "FFFFFF"
ROW_ALT = "F5F7FC"
TABLE_BORDER = "D1D9E6"
SLATE_TEXT = "262E3B"
PHOTO_SLOT = "F5F7FA"
SIG_BAR = "F7F7FA"
GREY_TEXT = "8C8C99"         # pdfPlacard rgb(0.55, 0.55, 0.6)
GREY_PLACEHOLDER = "9999A6"  # pdfPlacard rgb(0.6, 0.6, 0.65)
DRAFT_RED = "BF1A1A"         # the ES draft watermark
DRAFT_FILL = "FDECEC"
LINK_BLUE = "1155CC"

FONT_NAME = "Arial"
GRID_COLS = 12  # the placard's band grid; every band divides into these
MAX_ROW_BREAKS = 1026  # Excel's per-sheet ceiling on manual page breaks
COLUMN_WIDTH = 13      # character units; 12 of these span a landscape page
PHOTO_ROWS = 7         # rows a photo slot occupies
PHOTO_ROW_POINTS = 14  # height of each of those rows
BODY_ROW_POINTS = 12   # height of a wrapped body row
LINE_SPACING = 1.22    # Excel's line pitch as a multiple of the font size

# Excel sizes columns in character units and rows in points, but anchors
# images in pixels, so a slot's pixel box has to be derived from both.
PX_PER_CHAR, PX_CELL_PADDING, PX_PER_POINT = 7, 5, 4 / 3
COLUMN_PX = round(COLUMN_WIDTH * PX_PER_CHAR) + PX_CELL_PADDING
COLUMN_POINTS = COLUMN_PX * 0.75
CELL_PADDING_POINTS = 5

_CONTROL_CHARS = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")

# Arial, or the metrically identical Liberation Sans. Row heights are derived
# from measured text, so the font used to measure has to match the font Excel
# will wrap with; anything else mis-sizes every band.
_ARIAL_PATHS = (
    "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
    "/usr/share/fonts/truetype/liberation2/LiberationSans-Regular.ttf",
    "C:/Windows/Fonts/arial.ttf",
    "/Library/Fonts/Arial.ttf",
    "/System/Library/Fonts/Supplemental/Arial.ttf",
    "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
)
_MEASURE_SIZE = 200  # measure big, scale down — avoids integer-pixel rounding


@lru_cache(maxsize=None)
def _measuring_font():
    try:
        from PIL import ImageFont
    except ImportError:
        return None
    for path in _ARIAL_PATHS:
        if os.path.exists(path):
            try:
                return ImageFont.truetype(path, _MEASURE_SIZE)
            except OSError:
                continue
    return None


@lru_cache(maxsize=8192)
def _text_width(text: str, points: float) -> float:
    """Width of `text` in points at `points` size, in the placard's font."""
    font = _measuring_font()
    if font is None:
        # No font file to measure with: Arial's average lowercase advance is
        # close to half its em, which over-estimates narrow text and so errs
        # toward a taller row rather than a clipped one.
        return len(text) * points * 0.5
    return font.getlength(text) * points / _MEASURE_SIZE


def wrapped_lines(text: str, points: float, width_points: float) -> int:
    """How many lines Excel wraps `text` into within `width_points`."""
    if not text:
        return 1
    usable = max(width_points - CELL_PADDING_POINTS, points)
    total = 0
    for paragraph in str(text).split("\n"):
        words = paragraph.split()
        if not words:
            total += 1
            continue
        lines, current = 1, ""
        for word in words:
            candidate = f"{current} {word}".strip()
            if _text_width(candidate, points) <= usable or not current:
                current = candidate
            else:
                lines += 1
                current = word
                # A single word wider than the cell wraps again mid-word.
                while _text_width(current, points) > usable and len(current) > 1:
                    cut = max(1, int(len(current) * usable / _text_width(current, points)))
                    current = current[cut:]
                    lines += 1
        total += lines
    return total


def span_points(first_col: int, last_col: int) -> float:
    """Printable width of a merged column span, in points."""
    return (last_col - first_col + 1) * COLUMN_POINTS


def rows_for(text: str, points: float, first_col: int, last_col: int,
             row_points: float = BODY_ROW_POINTS, minimum: int = 1,
             maximum: int = 40) -> int:
    """Rows a wrapped block needs so Excel clips none of it.

    Excel never grows a merged cell to fit its contents — whatever overflows
    is simply not drawn. Every wrapped band therefore has to be sized from
    the text itself, or a placard quietly loses isolation steps.
    """
    lines = wrapped_lines(text, points, span_points(first_col, last_col))
    needed = -(-int(lines * points * LINE_SPACING + 2) // int(row_points))
    return max(minimum, min(maximum, needed))


def clean(value) -> str:
    """Text safe for an XLSX cell. Excel rejects most C0 control chars."""
    if value is None:
        return ""
    return _CONTROL_CHARS.sub("", str(value))


_BLANK_LINES = re.compile(r"\n[ \t]*\n+")


def placard_text(value) -> str:
    """Cell text for a placard band.

    Collapses runs of blank lines: most tag_description values are stored as
    a heading, an empty line, then the body, and that empty line costs a
    wrapped line in the narrowest column of every step on every page. Only
    whitespace is touched — the flat data sheets still carry the raw value.
    """
    return _BLANK_LINES.sub("\n", clean(value)).strip()


def cell_value(value):
    """Native type where Excel has one, so the data sheets stay sortable
    and filterable — a boolean written as the text "True" is neither."""
    if value is None:
        return None
    if isinstance(value, (bool, int, float)):
        return value
    return clean(value)


def energy_code(code: str | None) -> tuple[str, str, str, str, str]:
    raw = (code or "").strip().upper()
    return _BY_CODE.get(ENERGY_ALIASES.get(raw, raw)) or (
        raw or "?", raw or "Unknown", raw or "Desconocido", "888888", "FFFFFF"
    )


# ── Fetch ───────────────────────────────────────────────────────────────────
EQUIPMENT_COLUMNS = (
    "equipment_id,description,department,equipment_family,prefix,manufacturer,model,"
    "photo_status,has_equip_photo,has_iso_photo,needs_equip_photo,needs_iso_photo,"
    "needs_verification,verified,verified_date,verified_by,notes,notes_es,"
    "internal_notes,spanish_reviewed,equip_photo_url,iso_photo_url,placard_url,"
    "qr_token,readiness_status,last_pre_use_inspection_at,next_periodic_review_due_at,"
    "cad_drawing_tag,cad_sheet_ref,cad_grid_ref,facility_floor,flagged_for_review_at,"
    "flagged_for_review_by,flagged_for_review_via,flagged_for_review_note,"
    "equip_photo_provenance,iso_photo_provenance,iso_photo_is_placeholder,"
    "last_audit_verdict,last_audit_at,created_at,updated_at"
)
STEP_COLUMNS = (
    "equipment_id,step_number,sequence_order,step_type,energy_type,tryout_required,"
    "tag_description,isolation_procedure,method_of_verification,tag_description_es,"
    "isolation_procedure_es,method_of_verification_es,confidence"
)
PHOTO_MAX_PX = 900  # placard slots print ~370pt wide; 900px is generous at 300dpi


def _rest(base: str, key: str, path: str, params: dict) -> list[dict]:
    """One PostgREST GET, paged by the caller via Range params."""
    url = f"{base}/rest/v1/{path}?{urllib.parse.urlencode(params)}"
    request = urllib.request.Request(
        url,
        headers={
            "apikey": key,
            "Authorization": f"Bearer {key}",
            "Accept": "application/json",
        },
    )
    with urllib.request.urlopen(request, timeout=120) as response:
        return json.loads(response.read().decode("utf-8"))


def _paged(base: str, key: str, path: str, params: dict, page: int = 1000) -> list[dict]:
    rows: list[dict] = []
    while True:
        batch = _rest(base, key, path, {**params, "offset": len(rows), "limit": page})
        rows.extend(batch)
        if len(batch) < page:
            return rows


def cmd_fetch(args: argparse.Namespace) -> int:
    out = Path(args.out)
    (out / "photos").mkdir(parents=True, exist_ok=True)

    # Placard photos are public storage objects, so refreshing them needs no
    # credentials — only the cached equipment rows that name their URLs.
    if args.photos_only:
        equipment = json.loads((out / "equipment.json").read_text("utf-8"))
        print(f"{_download_photos(equipment, out / 'photos')} photos cached "
              f"-> {out / 'photos'}")
        return 0

    base = os.environ.get("SUPABASE_URL", "").rstrip("/")
    key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "")
    tenant_number = os.environ.get("SUPABASE_TENANT_NUMBER", "0001")
    if not base or not key:
        sys.exit("SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY must be set to fetch.")

    tenants = _rest(base, key, "tenants",
                    {"select": "id,name,tenant_number", "tenant_number": f"eq.{tenant_number}"})
    if not tenants:
        sys.exit(f"No tenant with tenant_number={tenant_number}.")
    tenant = tenants[0]

    equipment = _paged(base, key, "loto_equipment", {
        "select": EQUIPMENT_COLUMNS,
        "tenant_id": f"eq.{tenant['id']}",
        "decommissioned": "is.false",
        "order": "equipment_id.asc",
    })
    ids = {row["equipment_id"] for row in equipment}
    steps = [s for s in _paged(base, key, "loto_energy_steps", {
        "select": STEP_COLUMNS,
        "tenant_id": f"eq.{tenant['id']}",
        "order": "equipment_id.asc,sequence_order.asc,step_number.asc",
    }) if s["equipment_id"] in ids]

    (out / "equipment.json").write_text(json.dumps(equipment, ensure_ascii=False), "utf-8")
    (out / "steps.json").write_text(json.dumps(steps, ensure_ascii=False), "utf-8")
    (out / "tenant.json").write_text(json.dumps(tenant, ensure_ascii=False), "utf-8")
    print(f"{len(equipment)} equipment, {len(steps)} steps -> {out}")

    if args.no_photos:
        return 0
    downloaded = _download_photos(equipment, out / "photos")
    print(f"{downloaded} photos cached -> {out / 'photos'}")
    return 0


def _download_photos(equipment: list[dict], into: Path) -> int:
    from PIL import Image

    count = 0
    for row in equipment:
        for kind in ("equip", "iso"):
            url = row.get(f"{kind}_photo_url")
            if not url:
                continue
            target = into / f"{row['equipment_id'].replace('/', '_')}_{kind}.jpg"
            if target.exists():
                count += 1
                continue
            try:
                with urllib.request.urlopen(url, timeout=60) as response:
                    target.write_bytes(response.read())
                # Downscale in place: 996 full-resolution photos would push the
                # workbook past 300 MB, and the placard slot is ~370pt wide.
                with Image.open(target) as image:
                    image = image.convert("RGB")
                    image.thumbnail((PHOTO_MAX_PX, PHOTO_MAX_PX), Image.LANCZOS)
                    image.save(target, "JPEG", quality=82, optimize=True)
                count += 1
            except Exception as error:  # one bad photo must not sink the export
                print(f"  photo failed {row['equipment_id']} {kind}: {error}", file=sys.stderr)
                target.unlink(missing_ok=True)
    return count


# ── Build ───────────────────────────────────────────────────────────────────
@dataclass
class Style:
    """A cell's look, applied across a merged band in one call."""
    fill: str | None = None
    color: str = SLATE_TEXT
    size: float = 9
    bold: bool = False
    align: str = "left"
    valign: str = "center"
    wrap: bool = False
    border: bool = False


THIN = Side(style="thin", color=TABLE_BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


class PlacardSheet:
    """Writes placard pages onto one worksheet, band by band."""

    def __init__(self, worksheet, photos: Path):
        self.ws = worksheet
        self.photos = photos
        self.row = 1

    def merge(self, first_row: int, first_col: int, last_row: int, last_col: int) -> None:
        """Merge a band, skipping Worksheet.merge_cells.

        merge_cells tests the new range against every existing one before
        storing it, which is O(n) per call and quadratic over a workbook
        this size — around six minutes for 499 placards. The bands here are
        laid out row by row and never overlap, so the test only costs time.
        """
        self.ws.merged_cells.ranges.add(CellRange(
            min_col=first_col, min_row=first_row,
            max_col=last_col, max_row=last_row,
        ))

    def band(self, text, style: Style, height: float | None = None,
             span: tuple[int, int] = (1, GRID_COLS)) -> int:
        """Write one merged, styled band and advance to the next row."""
        row = self.row
        self.cell(row, span, text, style)
        if height is not None:
            self.ws.row_dimensions[row].height = height
        self.row += 1
        return row

    def cell(self, row: int, span: tuple[int, int], text, style: Style,
             merge: bool = True):
        """Write one single-row band. `merge=False` leaves the range to the
        caller, so a multi-row block does not also register a one-row merge
        over its own anchor — two ranges sharing a corner is a malformed
        workbook that Excel offers to repair on open."""
        first, last = span
        cell = self.ws.cell(row=row, column=first)
        cell.value = clean(text) if text != "" else None
        cell.font = Font(name=FONT_NAME, size=style.size, bold=style.bold, color=style.color)
        cell.alignment = Alignment(horizontal=style.align, vertical=style.valign,
                                   wrap_text=style.wrap)
        if style.fill:
            cell.fill = PatternFill("solid", fgColor=style.fill)
        if style.border:
            cell.border = BOX
        if last > first:
            if merge:
                self.merge(row, first, row, last)
            # Merged cells keep their own fill/border, so paint the tail too.
            for column in range(first + 1, last + 1):
                tail = self.ws.cell(row=row, column=column)
                if style.fill:
                    tail.fill = PatternFill("solid", fgColor=style.fill)
                if style.border:
                    tail.border = BOX
        return cell

    def block(self, rows: int, span: tuple[int, int], text, style: Style,
              start: int | None = None) -> int:
        """A band that spans several rows (wrapped body text, photo slots)."""
        row = self.row if start is None else start
        cell = self.cell(row, span, text, style, merge=False)
        if rows > 1 or span[1] > span[0]:
            self.merge(row, span[0], row + rows - 1, span[1])
        for r in range(row, row + rows):
            for c in range(span[0], span[1] + 1):
                tail = self.ws.cell(row=r, column=c)
                if style.fill:
                    tail.fill = PatternFill("solid", fgColor=style.fill)
                if style.border:
                    tail.border = BOX
        return cell


def _grouped_steps(steps: list[dict]) -> list[dict]:
    """Steps in the order the placard prints them: OSHA phase, then sequence."""
    ordered: list[dict] = []
    for phase in STEP_PHASE_ORDER:
        in_phase = [s for s in steps if s.get("step_type") == phase]
        in_phase.sort(key=lambda s: (s.get("sequence_order") or 0, s.get("step_number") or 0))
        ordered.extend(in_phase)
    known = set(STEP_PHASE_ORDER)
    leftover = [s for s in steps if s.get("step_type") not in known]
    leftover.sort(key=lambda s: (s.get("sequence_order") or 0, s.get("step_number") or 0))
    return ordered + leftover


def _pick(step: dict, field: str, language: str) -> str:
    """Spanish text where present, English as the placard's own fallback."""
    if language == "en":
        return placard_text(step.get(field))
    spanish = placard_text(step.get(f"{field}_es"))
    return spanish or placard_text(step.get(field))


def draw_placard(sheet: PlacardSheet, equipment: dict, steps: list[dict],
                 language: str, generated: str) -> None:
    """One placard page: the same bands, in the same order, as the PDF."""
    ws, is_en = sheet.ws, language == "en"
    half = GRID_COLS // 2

    # 1. Yellow title band. The QR caption shares the date line rather than
    # taking a row of its own — vertical space is the scarce resource here.
    top = sheet.row
    sheet.band(TEXT["title"][language],
               Style(fill=YELLOW_BAND, color="000000", size=14, bold=True, align="center"),
               height=24)
    date_label = f"{'Date' if is_en else 'Fecha'}: {generated}"
    row = sheet.row
    subtitle = sheet.cell(row, (1, GRID_COLS),
                          f"[{'EN' if is_en else 'ES'}]   {date_label}",
                          Style(fill=YELLOW_BAND, color="000000", size=8, align="center"))
    if equipment.get("qr_token"):
        target = f"https://soteriafield.app/qr/{equipment['qr_token']}"
        subtitle.value = (f"[{'EN' if is_en else 'ES'}]   {date_label}   ·   "
                          f"{'Scan for digital placard' if is_en else 'Escanee para placa digital'}")
        subtitle.hyperlink = target
    ws.row_dimensions[row].height = 12
    sheet.row += 1

    # 1b. Draft banner. The PDF stamps a rotated "BORRADOR — NO REVISADO"
    # watermark across any Spanish page whose translation is unreviewed. A
    # spreadsheet cannot rotate a watermark behind its cells, so the same
    # warning takes a band of its own — losing it would present an unreviewed
    # translation as an approved procedure.
    if not is_en and equipment.get("spanish_reviewed") is False:
        sheet.band("BORRADOR — NO REVISADO  ·  traducción sin revisar",
                   Style(fill=DRAFT_FILL, color=DRAFT_RED, size=9, bold=True,
                         align="center"), height=13)

    # 2. Blue equipment bar — description left, department right.
    row = sheet.row
    sheet.cell(row, (1, 8), f"{TEXT['equipment_label'][language]} {clean(equipment.get('description'))}",
               Style(fill=BLUE_BAR, color=NAVY_HEADER, size=10, bold=True))
    sheet.cell(row, (9, GRID_COLS), clean(equipment.get("department")),
               Style(fill=BLUE_BAR, color=NAVY_HEADER, size=10, bold=True, align="right"))
    ws.row_dimensions[row].height = 16
    sheet.row += 1

    # 3. Red warning block. The PDF prints only the first wrapped line of the
    # note; here the whole note is shown, because the band is the one place a
    # machine-specific hazard is written down.
    sheet.band(TEXT["warning_header"][language],
               Style(fill=RED_BLOCK, color=WHITE, size=9.5, bold=True, align="center"),
               height=14)
    notes = placard_text(equipment.get("notes") if is_en else equipment.get("notes_es"))
    warning = notes or TEXT["warning_fallback"][language]
    warning_rows = rows_for(warning, 8, 1, GRID_COLS, minimum=1, maximum=4)
    sheet.block(warning_rows, (1, GRID_COLS), warning,
                Style(fill=RED_BLOCK, color=WHITE, size=8, align="center", wrap=True))
    for r in range(sheet.row, sheet.row + warning_rows):
        ws.row_dimensions[r].height = BODY_ROW_POINTS
    sheet.row += warning_rows

    # 4. Purpose | Lockout application process.
    row = sheet.row
    sheet.cell(row, (1, 7), TEXT["purpose_header"][language],
               Style(color=NAVY_HEADER, size=9, bold=True))
    sheet.cell(row, (8, GRID_COLS), TEXT["application_header"][language],
               Style(color=NAVY_HEADER, size=9, bold=True))
    ws.row_dimensions[row].height = 12
    sheet.row += 1

    purpose = TEXT["purpose_body"][language]
    application = "\n".join(f"{i}. {s}" for i, s in
                            enumerate(TEXT["application_steps"][language], 1))
    body_rows = max(rows_for(purpose, 8, 1, 7),
                    rows_for(application, 8, 8, GRID_COLS))
    start = sheet.row
    sheet.block(body_rows, (1, 7), purpose,
                Style(size=8, wrap=True, valign="top"), start=start)
    sheet.block(body_rows, (8, GRID_COLS), application,
                Style(size=8, wrap=True, valign="top"), start=start)
    for r in range(start, start + body_rows):
        ws.row_dimensions[r].height = BODY_ROW_POINTS
    sheet.row = start + body_rows

    # 5. Colour codes — red strip over a 6x2 grid of the 12 real codes.
    sheet.band(TEXT["color_codes_header"][language],
               Style(fill=RED_BLOCK, color=WHITE, size=9, bold=True, align="center"), height=14)
    codes = [c for c in ENERGY_CODES if c[0] != "N"]
    per_row = 6
    for chunk_start in range(0, len(codes), per_row):
        row = sheet.row
        for offset, (code, label_en, label_es, fill, text_color) in enumerate(
                codes[chunk_start:chunk_start + per_row]):
            column = 1 + offset * 2
            sheet.cell(row, (column, column + 1),
                       f"{code} = {label_en if is_en else label_es}",
                       Style(fill=fill, color=text_color, size=7, bold=True,
                             align="center", border=True))
        ws.row_dimensions[row].height = 12
        sheet.row += 1

    # 6. Navy section header.
    sheet.band(TEXT["section_header"][language],
               Style(fill=NAVY_HEADER, color=WHITE, size=9, bold=True), height=14)

    # 7. Photo slots, side by side.
    caption_en, caption_es = TEXT["photo_captions"][language]
    photo_rows = PHOTO_ROWS
    start = sheet.row
    for index, (kind, caption) in enumerate((("equip", caption_en), ("iso", caption_es))):
        span = (1, half) if index == 0 else (half + 1, GRID_COLS)
        path = sheet.photos / f"{equipment['equipment_id'].replace('/', '_')}_{kind}.jpg"
        url = equipment.get(f"{kind}_photo_url")
        if path.exists():
            sheet.block(photo_rows, span, "", Style(fill=PHOTO_SLOT, border=True), start=start)
            _embed(ws, path, start, span[0], photo_rows)
        else:
            # The URL goes on the hyperlink, never into the cell text: it is one
            # unbreakable token, so Excel cannot wrap it and it would sprawl
            # across the neighbouring slot.
            label = TEXT["no_photo"][language]
            if url:
                label += "\n" + ("click to open the photo" if is_en
                                 else "haga clic para abrir la foto")
            cell = sheet.block(photo_rows, span, label,
                               Style(fill=PHOTO_SLOT, color=LINK_BLUE if url else GREY_PLACEHOLDER,
                                     size=8, align="center", wrap=True, border=True),
                               start=start)
            if url:
                cell.hyperlink = url
        sheet.cell(start + photo_rows, span, caption,
                   Style(fill=PHOTO_SLOT, color=NAVY_HEADER, size=7.5, bold=True, border=True))
    for r in range(start, start + photo_rows):
        ws.row_dimensions[r].height = PHOTO_ROW_POINTS
    ws.row_dimensions[start + photo_rows].height = 11
    sheet.row = start + photo_rows + 1

    # 8. Energy steps table.
    header = TEXT["table_headers"][language]
    row = sheet.row
    header_style = Style(fill=NAVY_HEADER, color=WHITE, size=8, bold=True, border=True)
    sheet.cell(row, (1, 2), header[0], header_style)
    sheet.cell(row, (3, 7), header[1], header_style)
    sheet.cell(row, (8, GRID_COLS), header[2], header_style)
    ws.row_dimensions[row].height = 14
    sheet.row += 1

    ordered = _grouped_steps(steps)
    if not ordered:
        sheet.band(TEXT["no_steps"][language],
                   Style(color=GREY_TEXT, size=9, align="center", border=True), height=20)
    for index, step in enumerate(ordered):
        code, label_en, label_es, fill, text_color = energy_code(step.get("energy_type"))
        tag = _pick(step, "tag_description", language)
        procedure = _pick(step, "isolation_procedure", language)
        verification = _pick(step, "method_of_verification", language)
        phase = PHASE_LABELS[language].get(step.get("step_type"), clean(step.get("step_type")))
        alt = ROW_ALT if index % 2 else None

        badge = f"{code} — {label_en if is_en else label_es}\n{phase}"
        if tag:
            badge += f"\n{tag}"
        # Sized to the tallest of the three columns — the badge carries three
        # lines of its own and used to be the one that clipped.
        rows_needed = max(
            rows_for(badge, 7, 1, 2, minimum=2),
            rows_for(procedure, 7.5, 3, 7, minimum=2),
            rows_for(verification, 7.5, 8, GRID_COLS, minimum=2),
        )
        start = sheet.row
        sheet.block(rows_needed, (1, 2), badge,
                    Style(fill=fill, color=text_color, size=7, bold=True,
                          wrap=True, valign="top", border=True), start=start)
        sheet.block(rows_needed, (3, 7), procedure,
                    Style(fill=alt, size=7.5, wrap=True, valign="top", border=True), start=start)
        sheet.block(rows_needed, (8, GRID_COLS), verification,
                    Style(fill=alt, size=7.5, wrap=True, valign="top", border=True), start=start)
        for r in range(start, start + rows_needed):
            ws.row_dimensions[r].height = BODY_ROW_POINTS
        sheet.row = start + rows_needed

    # 9. Lockout removal process — red strip over two columns.
    sheet.band(TEXT["removal_header"][language],
               Style(fill=RED_BLOCK, color=WHITE, size=9, bold=True, align="center"), height=13)
    removal = TEXT["removal_steps"][language]
    split = -(-len(removal) // 2)
    left = "\n".join(f"{i}. {s}" for i, s in enumerate(removal[:split], 1))
    right = "\n".join(f"{i}. {s}" for i, s in enumerate(removal[split:], split + 1))
    removal_rows = max(rows_for(left, 7.5, 1, 6), rows_for(right, 7.5, 7, GRID_COLS))
    start = sheet.row
    sheet.block(removal_rows, (1, 6), left,
                Style(size=7.5, wrap=True, valign="top"), start=start)
    sheet.block(removal_rows, (7, GRID_COLS), right,
                Style(size=7.5, wrap=True, valign="top"), start=start)
    for r in range(start, start + removal_rows):
        ws.row_dimensions[r].height = BODY_ROW_POINTS
    sheet.row = start + removal_rows

    # 10. Print note.
    sheet.band(TEXT["print_note"][language], Style(color=GREY_TEXT, size=6.5), height=9)

    # 11. Signature bar.
    row = sheet.row
    for index, label in enumerate(TEXT["signature"][language]):
        column = 1 + index * 3
        sheet.cell(row, (column, column + 2), label,
                   Style(fill=SIG_BAR, color=NAVY_HEADER, size=8, bold=True, border=True))
    ws.row_dimensions[row].height = 18
    sheet.row += 1

    # A blank spacer separates placards on screen. It has to be added before
    # the break, or it lands after it and every printed page opens with an
    # empty row.
    sheet.row += 1
    # One placard per printed page. Excel stores at most 1026 manual row
    # breaks per sheet and repairs the file if it finds more; past that the
    # pages flow together, which cmd_build reports rather than hiding.
    if len(ws.row_breaks.brk) < MAX_ROW_BREAKS:
        ws.row_breaks.append(Break(id=sheet.row - 1))
    return top


def _embed(ws, path: Path, row: int, column: int, rows: int) -> None:
    """Place a cached photo inside its slot, scaled to fit and centred."""
    from PIL import Image

    column_px = round(COLUMN_WIDTH * PX_PER_CHAR) + PX_CELL_PADDING
    slot_width = (GRID_COLS // 2) * column_px - 6      # 6px of breathing room
    slot_height = round(rows * PHOTO_ROW_POINTS * PX_PER_POINT) - 6
    with Image.open(path) as probe:
        width, height = probe.size
    scale = min(slot_width / width, slot_height / height, 1.0)
    drawn_width, drawn_height = int(width * scale), int(height * scale)

    image = XLImage(str(path))
    image.width, image.height = drawn_width, drawn_height
    # A plain cell anchor pins the image to the slot's top-left corner; the
    # offsets recentre it the way the PDF centres its photo slots.
    image.anchor = OneCellAnchor(
        _from=AnchorMarker(
            col=column - 1, colOff=pixels_to_EMU((slot_width - drawn_width) // 2),
            row=row - 1, rowOff=pixels_to_EMU((slot_height - drawn_height) // 2),
        ),
        ext=XDRPositiveSize2D(pixels_to_EMU(drawn_width), pixels_to_EMU(drawn_height)),
    )
    ws.add_image(image)


def _flat_sheet(ws, rows: list[dict], columns: list[str]) -> None:
    """A plain filterable table — the data behind the placard pages."""
    ws.append(columns)
    for cell in ws[1]:
        cell.font = Font(name=FONT_NAME, size=10, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=NAVY_HEADER)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30
    for row in rows:
        ws.append([cell_value(row.get(column)) for column in columns])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(rows) + 1}"
    for index, column in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(index)].width = min(
            max(12, len(column) + 2), 46
        )


def cmd_build(args: argparse.Namespace) -> int:
    data = Path(args.data)
    equipment = json.loads((data / "equipment.json").read_text("utf-8"))
    steps = json.loads((data / "steps.json").read_text("utf-8"))
    photos = data / "photos"
    languages = ["en", "es"] if args.language == "both" else [args.language]

    by_equipment: dict[str, list[dict]] = {}
    for step in steps:
        by_equipment.setdefault(step["equipment_id"], []).append(step)
    equipment.sort(key=lambda row: row["equipment_id"])

    workbook = Workbook()
    workbook.remove(workbook.active)
    # Created in reading order up front, then filled — the reader meets the
    # Read Me and Index before the placards themselves.
    readme, index_sheet, placards, equipment_sheet, steps_sheet, summary = (
        workbook.create_sheet(name) for name in
        ("Read Me", "Index", "Placards", "Equipment", "Energy Steps", "Summary")
    )

    for column in range(1, GRID_COLS + 1):
        placards.column_dimensions[get_column_letter(column)].width = COLUMN_WIDTH
    placards.page_setup.orientation = "landscape"
    placards.page_setup.paperSize = placards.PAPERSIZE_LETTER
    placards.page_setup.fitToWidth = 1
    placards.page_setup.fitToHeight = 0
    placards.sheet_properties.pageSetUpPr.fitToPage = True
    placards.print_options.horizontalCentered = True
    # Narrow margins. openpyxl defaults to an inch top and bottom, which costs
    # 144pt of a 612pt page — more than the photo band and the steps table
    # header put together, and the difference between one page and two.
    placards.page_margins = PageMargins(left=0.3, right=0.3, top=0.3, bottom=0.3,
                                        header=0.15, footer=0.15)

    sheet = PlacardSheet(placards, photos)
    index: list[tuple[str, str, str, int]] = []
    for row in equipment:
        for language in languages:
            top = draw_placard(sheet, row, by_equipment.get(row["equipment_id"], []),
                               language, args.date)
            index.append((row["equipment_id"], clean(row.get("department")), language, top))

    _flat_sheet(equipment_sheet, equipment,
                list(equipment[0].keys()) if equipment else [])
    _flat_sheet(steps_sheet, steps, list(steps[0].keys()) if steps else [])
    _index_sheet(index_sheet, index)
    _summary_sheet(summary, equipment, len(steps), len(index))
    _readme_sheet(readme, equipment, steps, photos, args,
                  fit=_page_fit(placards, sorted(b.id for b in placards.row_breaks.brk)),
                  pages=len(index))

    workbook.save(args.out)
    print(f"{len(equipment)} placards x {len(languages)} language(s) -> {args.out}")
    if len(index) > MAX_ROW_BREAKS:
        print(f"  note: Excel caps a sheet at {MAX_ROW_BREAKS} manual page breaks, so the "
              f"last {len(index) - MAX_ROW_BREAKS} placard(s) share a page with the one "
              f"before. Export a single language, or split the run, to page them all.",
              file=sys.stderr)
    return 0


def _index_sheet(ws, index: list[tuple[str, str, str, int]]) -> None:
    ws.append(["Equipment ID", "Department", "Language", "Go to placard"])
    for cell in ws[1]:
        cell.font = Font(name=FONT_NAME, size=10, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=NAVY_HEADER)
    for equipment_id, department, language, row in index:
        ws.append([equipment_id, department, language.upper(), None])
        link = ws.cell(row=ws.max_row, column=4)
        link.value = "Open"
        link.hyperlink = f"#'Placards'!A{row}"
        link.font = Font(name=FONT_NAME, size=10, color="1155CC", underline="single")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:D{ws.max_row}"
    for column, width in zip("ABCD", (22, 28, 10, 14)):
        ws.column_dimensions[column].width = width


def _summary_sheet(ws, equipment: list[dict], step_count: int, page_count: int) -> None:
    """Per-department counts, computed by formula against the Equipment sheet."""
    ws["A1"] = "Placard export summary"
    ws["A1"].font = Font(name=FONT_NAME, size=13, bold=True, color=NAVY_HEADER)

    last = len(equipment) + 1
    department_column = (list(equipment[0].keys()).index("department") + 1) if equipment else 3
    status_column = (list(equipment[0].keys()).index("photo_status") + 1) if equipment else 8
    verified_column = (list(equipment[0].keys()).index("verified") + 1) if equipment else 14
    dept_range = f"Equipment!${get_column_letter(department_column)}$2:${get_column_letter(department_column)}${last}"
    status_range = f"Equipment!${get_column_letter(status_column)}$2:${get_column_letter(status_column)}${last}"
    verified_range = f"Equipment!${get_column_letter(verified_column)}$2:${get_column_letter(verified_column)}${last}"

    ws["A3"], ws["B3"] = "Placards (active equipment)", len(equipment)
    ws["A4"], ws["B4"] = "Energy isolation steps", step_count
    ws["A5"], ws["B5"] = "Rendered placard pages", page_count
    for row in (3, 4, 5):
        ws[f"A{row}"].font = Font(name=FONT_NAME, size=10)
        ws[f"B{row}"].font = Font(name=FONT_NAME, size=10, bold=True)

    headers = ["Department", "Placards", "Photos complete", "Photos partial",
               "Photos missing", "Verified"]
    ws.append([])
    ws.append(headers)
    header_row = ws.max_row
    for cell in ws[header_row]:
        cell.font = Font(name=FONT_NAME, size=10, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=NAVY_HEADER)

    for department in sorted({clean(row.get("department")) for row in equipment}):
        ws.append([department])
        row = ws.max_row
        ws.cell(row=row, column=2).value = f'=COUNTIF({dept_range},$A{row})'
        for column, status in ((3, "complete"), (4, "partial"), (5, "missing")):
            ws.cell(row=row, column=column).value = (
                f'=COUNTIFS({dept_range},$A{row},{status_range},"{status}")'
            )
        ws.cell(row=row, column=6).value = (
            f"=COUNTIFS({dept_range},$A{row},{verified_range},TRUE)"
        )
        for column in range(1, 7):
            ws.cell(row=row, column=column).font = Font(name=FONT_NAME, size=10)

    total = ws.max_row + 1
    ws.cell(row=total, column=1).value = "Total"
    for column in range(2, 7):
        letter = get_column_letter(column)
        ws.cell(row=total, column=column).value = f"=SUM({letter}{header_row + 1}:{letter}{total - 1})"
    for column in range(1, 7):
        ws.cell(row=total, column=column).font = Font(name=FONT_NAME, size=10, bold=True)

    ws.column_dimensions["A"].width = 30
    for column in "BCDEF":
        ws.column_dimensions[column].width = 16
    ws.freeze_panes = f"A{header_row + 1}"


def _page_fit(ws, breaks: list[int]) -> tuple[int, int]:
    """How many placards print on one page, and how many pages the rest take.

    Fit-to-width scales the sheet down by the ratio of printable width to
    sheet width, and that same factor shrinks the rows, so the comparison
    has to be made after scaling.
    """
    margins = ws.page_margins
    printable_w = 792 - (margins.left + margins.right) * 72
    printable_h = 612 - (margins.top + margins.bottom) * 72
    sheet_w = GRID_COLS * COLUMN_POINTS
    scale = min(1.0, printable_w / sheet_w)

    starts = [1] + [b + 1 for b in breaks[:-1]]
    on_one_page, most_pages = 0, 1
    for start, end in zip(starts, breaks):
        height = sum(ws.row_dimensions[r].height or 15.0 for r in range(start, end + 1))
        pages = max(1, -(-int(height * scale) // int(printable_h)))
        on_one_page += pages == 1
        most_pages = max(most_pages, pages)
    return on_one_page, most_pages


def _readme_sheet(ws, equipment: list[dict], steps: list[dict], photos: Path,
                  args: argparse.Namespace, fit: tuple[int, int] = (0, 1),
                  pages: int = 0) -> None:
    cached = len(list(photos.glob("*.jpg"))) if photos.exists() else 0
    expected = sum(1 for row in equipment for kind in ("equip", "iso")
                   if row.get(f"{kind}_photo_url"))
    on_one_page, most_pages = fit
    lines = [
        ("Soteria Field — LOTO placard export", 15, True, NAVY_HEADER),
        ("", 10, False, SLATE_TEXT),
        (f"Generated: {args.date}", 10, False, SLATE_TEXT),
        (f"Placards (active, non-decommissioned equipment): {len(equipment)}", 10, False, SLATE_TEXT),
        (f"Energy isolation steps: {len(steps)}", 10, False, SLATE_TEXT),
        (f"Photos embedded: {cached} of {expected} referenced", 10, False, SLATE_TEXT),
        ("", 10, False, SLATE_TEXT),
        ("Sheets", 12, True, NAVY_HEADER),
        ("Index — every placard page, with a link into the Placards sheet.", 10, False, SLATE_TEXT),
        ("Placards — one printable page per placard per language, laid out to "
         "match the PDF: title band, equipment bar, warning block, purpose and "
         "application process, colour codes, photos, energy-steps table, removal "
         "process, signature bar.", 10, False, SLATE_TEXT),
        ("Equipment — the loto_equipment row behind each placard, filterable.", 10, False, SLATE_TEXT),
        ("Energy Steps — every loto_energy_steps row, filterable.", 10, False, SLATE_TEXT),
        ("Summary — per-department counts, computed by formula.", 10, False, SLATE_TEXT),
        ("", 10, False, SLATE_TEXT),
        ("Printing", 12, True, NAVY_HEADER),
        ("The Placards sheet is landscape, fit-to-width, with a page break after "
         "each placard, so no placard ever shares a sheet with another.", 10, False, SLATE_TEXT),
        (f"{on_one_page} of {pages} placard pages fit a single Letter sheet; the rest "
         f"run to {most_pages if most_pages > 1 else 2} sheets. That is the deliberate "
         "difference from the PDF: the PDF fits every placard on one page by dropping "
         "the isolation-step rows that overflow, and a silently shortened isolation "
         "procedure is a safety defect, so this export keeps the full text and takes "
         "the extra sheet instead. Machines with the most energy sources are the ones "
         "that run long.", 10, False, SLATE_TEXT),
        ("", 10, False, SLATE_TEXT),
        ("Regenerating", 12, True, NAVY_HEADER),
        ("  python scripts/export_placards_xlsx.py fetch --out data/", 10, False, SLATE_TEXT),
        ("  python scripts/export_placards_xlsx.py build --data data/ --out placards.xlsx",
         10, False, SLATE_TEXT),
        ("fetch needs SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY and downloads the "
         "photos; build works offline from whatever fetch cached.", 10, False, SLATE_TEXT),
    ]
    for text, size, bold, color in lines:
        ws.append([text])
        cell = ws.cell(row=ws.max_row, column=1)
        cell.font = Font(name=FONT_NAME, size=size, bold=bold, color=color)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.column_dimensions["A"].width = 118


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__,
                                     formatter_class=argparse.RawDescriptionHelpFormatter)
    sub = parser.add_subparsers(dest="command", required=True)

    fetch = sub.add_parser("fetch", help="cache equipment, steps, and photos from Supabase")
    fetch.add_argument("--out", default="placard-data", help="directory to write the cache into")
    fetch.add_argument("--no-photos", action="store_true", help="skip photo download")
    fetch.add_argument("--photos-only", action="store_true",
                       help="download only the photos named by a cached equipment.json "
                            "(public objects; needs no credentials)")
    fetch.set_defaults(func=cmd_fetch)

    build = sub.add_parser("build", help="render the workbook from a cache directory")
    build.add_argument("--data", default="placard-data", help="directory fetch wrote")
    build.add_argument("--out", default="loto-placards.xlsx", help="workbook to write")
    build.add_argument("--language", choices=("en", "es", "both"), default="both")
    build.add_argument("--date", default="", help="date printed on each placard")
    build.set_defaults(func=cmd_build)

    args = parser.parse_args()
    if getattr(args, "date", None) == "":
        from datetime import date
        args.date = date.today().isoformat()
    return args.func(args)


if __name__ == "__main__":
    raise SystemExit(main())
