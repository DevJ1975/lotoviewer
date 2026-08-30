#!/usr/bin/env python3
"""Find text that Excel will silently clip in a placard workbook.

Excel never grows a merged cell to fit its contents: whatever overflows the
merged height is simply not drawn, with no error and no visual cue. On a
LOTO placard that means a missing line of isolation procedure, so this runs
over every text cell in the Placards sheet, re-wraps it, and compares the
result against the height actually allocated.

    python scripts/check_placards_xlsx.py placards.xlsx

Exits non-zero if anything is clipped.

Nothing here is imported from export_placards_xlsx, deliberately. An earlier
version of this check shared the exporter's column-width constant and its
wrapping code, so it could only ever confirm that the exporter agreed with
itself — and 753 clipped cells passed it. The geometry below is derived
from the workbook and from the OOXML spec instead.
"""

from __future__ import annotations

import sys
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

MDW = 7                      # max digit width, Calibri/Arial 11px grid
CELL_INSET_POINTS = 5 * 0.75  # Excel's horizontal text inset, 5px
LINE_SPACING = 1.31          # Excel autofit for Arial: 1.275x @10pt, 1.312x @12pt
DEFAULT_ROW_POINTS = 15.0
MEASURE_SIZE = 200           # measure large, scale down, to dodge pixel rounding

ARIAL_PATHS = {
    False: (
        "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
        "/usr/share/fonts/truetype/liberation2/LiberationSans-Regular.ttf",
        "C:/Windows/Fonts/arial.ttf",
        "/Library/Fonts/Arial.ttf",
        "/System/Library/Fonts/Supplemental/Arial.ttf",
    ),
    True: (
        "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf",
        "/usr/share/fonts/truetype/liberation2/LiberationSans-Bold.ttf",
        "C:/Windows/Fonts/arialbd.ttf",
        "/Library/Fonts/Arial Bold.ttf",
        "/System/Library/Fonts/Supplemental/Arial Bold.ttf",
    ),
}


def column_points(stored_width: float) -> float:
    """Rendered width of a column, in points, from its OOXML `width`.

    ECMA-376 18.3.1.13. The stored width already carries the five pixels of
    cell padding, so adding them again over-states every column by ~5%.
    """
    return int(((256 * stored_width + int(128 / MDW)) / 256) * MDW) * 0.75


def load_fonts():
    """Regular and bold. Bold is materially wider, and the energy badges and
    every band header are bold — measuring them with the regular face
    under-states their width."""
    from PIL import ImageFont

    faces = {}
    for bold, paths in ARIAL_PATHS.items():
        for path in paths:
            if Path(path).exists():
                faces[bold] = ImageFont.truetype(path, MEASURE_SIZE)
                break
    if False not in faces:
        sys.exit("No Arial-metric font found; cannot measure text.")
    faces.setdefault(True, faces[False])
    return faces


def span_label(first_col: int, last_col: int) -> str:
    return f"{get_column_letter(first_col)}:{get_column_letter(last_col)}"


def main() -> int:
    if len(sys.argv) < 2:
        sys.exit(__doc__)
    workbook = load_workbook(sys.argv[1])
    if "Placards" not in workbook.sheetnames:
        sys.exit("No 'Placards' sheet in this workbook.")
    ws = workbook["Placards"]
    fonts = load_fonts()

    width_of_column = {}
    for index in range(1, ws.max_column + 1):
        stored = ws.column_dimensions[get_column_letter(index)].width or 8.43
        width_of_column[index] = column_points(stored)

    def text_width(text: str, points: float, bold: bool = False) -> float:
        return fonts[bold].getlength(text) * points / MEASURE_SIZE

    def wrapped_lines(text: str, points: float, available: float,
                      bold: bool = False) -> int:
        usable = max(available - CELL_INSET_POINTS, points)
        total = 0
        for paragraph in str(text).split("\n"):
            words = paragraph.split()
            if not words:
                total += 1
                continue
            lines, current = 1, ""
            for word in words:
                candidate = f"{current} {word}".strip()
                if text_width(candidate, points, bold) <= usable or not current:
                    current = candidate
                else:
                    lines += 1
                    current = word
                while text_width(current, points, bold) > usable and len(current) > 1:
                    cut = max(1, int(len(current) * usable / text_width(current, points, bold)))
                    current = current[cut:]
                    lines += 1
            total += lines
        return total

    anchors = {(m.min_row, m.min_col): m for m in ws.merged_cells.ranges}
    overflows, checked = [], 0

    for row in ws.iter_rows():
        for cell in row:
            if not isinstance(cell.value, str) or not cell.value or cell.value.startswith("="):
                continue
            merged = anchors.get((cell.row, cell.column))
            wraps = bool(cell.alignment and cell.alignment.wrap_text)
            if not wraps and merged is None:
                # A lone unwrapped cell spills into empty neighbours, which is
                # cosmetic. A MERGED one does not spill — Excel clips it at the
                # merge boundary — so those are checked below.
                continue
            last_col = merged.max_col if merged else cell.column
            last_row = merged.max_row if merged else cell.row

            available = sum(width_of_column.get(c, 0)
                            for c in range(cell.column, last_col + 1))
            height = sum(ws.row_dimensions[r].height or DEFAULT_ROW_POINTS
                         for r in range(cell.row, last_row + 1))
            size = cell.font.size or 11
            bold = bool(cell.font.bold)

            checked += 1
            if wraps:
                needed = wrapped_lines(cell.value, size, available, bold) * size * LINE_SPACING
                short = needed - height
                if short > 0.5:
                    overflows.append((cell.row, span_label(cell.column, last_col),
                                      short, height, cell.value))
            else:
                # One line, clipped horizontally at the merge boundary.
                widest = max(text_width(line, size, bold)
                             for line in str(cell.value).split("\n"))
                short = widest - (available - CELL_INSET_POINTS)
                if short > 0.5 or size * LINE_SPACING > height + 0.5:
                    overflows.append((cell.row, span_label(cell.column, last_col),
                                      short, available, cell.value))

    print(f"checked {checked} cells that Excel clips rather than spills")
    print(f"clipped: {len(overflows)}")
    for row, span, short, have, text in sorted(overflows, key=lambda o: -o[2])[:15]:
        print(f"  row {row:6d} {span:>7}  short {short:6.1f}pt of {have:6.1f}pt  "
              f"{text[:70].replace(chr(10), ' / ')}")
    if overflows:
        print("\nby column span:")
        for span, count in Counter(o[1] for o in overflows).most_common():
            print(f"  {span:>7}: {count}")
    return 1 if overflows else 0


if __name__ == "__main__":
    raise SystemExit(main())
